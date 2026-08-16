import pytest
import openpyxl
from bot.export import create_excel_report
from bot.settlement import calculate_settlement


def test_excel_export_generation():
    families = [
        {"id": 1, "name": "Family A", "weight": 2.0},
        {"id": 2, "name": "Family B", "weight": 1.5},
        {"id": 3, "name": "Family C", "weight": 1.0},
    ]
    meals = [
        {"id": 10, "meal_number": 1, "name": "Breakfast"},
        {"id": 11, "meal_number": 2, "name": "BBQ"},
    ]
    expenses = [
        {"id": 20, "description": "Firewood", "amount": 30.0, "family_id": 1, "family_name": "Family A"},
    ]
    meal_contributions = {
        10: [{"family_id": 1, "amount": 40.0}],
        11: [{"family_id": 2, "amount": 90.0}],
    }
    meal_absences = {
        11: [3],  # Family C skipped BBQ
    }

    excel_buf = create_excel_report(
        trip_name="Test Camp",
        families=families,
        meals=meals,
        expenses=expenses,
        meal_contributions=meal_contributions,
        meal_absences=meal_absences,
    )

    wb = openpyxl.load_workbook(excel_buf, data_only=False)
    ws = wb["Expense Report"]

    # Verify Header
    assert ws["A1"].value == "BachzTab — Test Camp Expense Report"
    assert ws["A4"].value == "Item / Expense"
    assert ws["E4"].value == "Family A (w=2.0)"
    assert ws["F4"].value == "Family B (w=1.5)"
    assert ws["G4"].value == "Family C (w=1.0)"

    # Verify Math against calculate_settlement
    res = calculate_settlement(
        families=families,
        meals=meals,
        meal_contributions=meal_contributions,
        meal_absences=meal_absences,
        shared_expenses=expenses,
    )

    # Check Total Spent ($160.00)
    assert res.total_spent == 160.0
