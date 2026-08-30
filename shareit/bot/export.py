"""Excel export generator for BachzTab trip expense breakdown."""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_report(
    trip_name: str,
    families: list[dict],
    meals: list[dict],
    expenses: list[dict],
    meal_contributions: dict[int, list[dict]] | None = None,
    meal_absences: dict[int, list[int]] | None = None,
    meal_groupings: dict[int, list[dict]] | None = None,
    expense_groupings: dict[int, list[dict]] | None = None,
    group_title: str | None = None,
    settlement_result: object | None = None,
) -> io.BytesIO:
    """
    Generate an itemized Excel workbook (.xlsx) containing all meals and expenses,
    cost shares per family, notes on absences, and summary rows (Paid, Owed, Net Balance).
    """
    if settlement_result is None:
        from bot.settlement import calculate_settlement
        settlement_result = calculate_settlement(
            families=families,
            meals=meals,
            meal_contributions=meal_contributions or {},
            meal_absences=meal_absences or {},
            shared_expenses=expenses,
            meal_groupings=meal_groupings,
            expense_groupings=expense_groupings,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    ws.views.sheetView[0].showGridLines = True

    # 1. Title Banner
    report_title = group_title or trip_name
    ws.merge_cells("A1:D1")
    ws["A1"] = f"BachzTab — {report_title} Expense Report"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(vertical="center")

    ws["A2"] = f"Generated for {len(families)} families. Currency: USD ($)"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

    # 2. Table Headers (Row 4)
    family_ids = [f["id"] for f in families]
    family_names = {f["id"]: f["name"] for f in families}

    headers = ["Item / Expense", "Type", "Payer", "Total ($)"]
    for f in families:
        headers.append(f"{f['name']} (w={f['weight']})")
    headers.append("Notes / Absences")

    ws.append([])  # Row 3 empty
    ws.append(headers)  # Row 4 headers

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 3. Data Rows directly from settlement_result
    current_row = 5
    paid_totals = settlement_result.paid_by_family

    for item in settlement_result.item_shares:
        if item.payer_family_ids:
            payer_str = ", ".join(family_names.get(fid, "Family") for fid in item.payer_family_ids)
        else:
            payer_str = "None"

        type_str = "Meal/Event" if item.item_type == "meal" else "General Expense"

        row_data = [
            item.item_name,
            type_str,
            payer_str,
            item.total_amount,
        ]
        for fid in family_ids:
            row_data.append(item.family_shares.get(fid, 0.0))
        row_data.append(item.notes or "-")

        ws.append(row_data)

        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.border = thin_border
            if col_idx >= 4 and col_idx < 4 + len(family_ids) + 1:
                c.number_format = "$#,##0.00"
                c.alignment = Alignment(horizontal="right")
        current_row += 1

    # 4. Summary Rows
    data_start_row = 5
    data_end_row = current_row - 1 if current_row > 5 else 5

    # Summary Row 1: Total Owed (Cost Share)
    row_owed = ["Total Owed (Cost Share)", "-", "-", f"=SUM(D{data_start_row}:D{data_end_row})"]
    for idx, fid in enumerate(family_ids):
        col_letter = get_column_letter(5 + idx)
        row_owed.append(f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
    row_owed.append("-")

    ws.append(row_owed)
    owed_row_idx = current_row
    owed_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thick_bottom = Border(
        top=Side(style="thin", color="1F4E78"),
        bottom=Side(style="double", color="1F4E78"),
    )

    for col_idx in range(1, len(row_owed) + 1):
        c = ws.cell(row=owed_row_idx, column=col_idx)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = owed_fill
        c.border = thick_bottom
        if col_idx >= 4 and col_idx < 4 + len(family_ids) + 1:
            c.number_format = "$#,##0.00"
            c.alignment = Alignment(horizontal="right")
    current_row += 1

    # Summary Row 2: Total Paid
    row_paid = ["Total Paid by Family", "-", "-", f"=SUM(D{data_start_row}:D{data_end_row})"]
    for fid in family_ids:
        row_paid.append(paid_totals[fid])
    row_paid.append("-")

    ws.append(row_paid)
    paid_row_idx = current_row
    paid_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col_idx in range(1, len(row_paid) + 1):
        c = ws.cell(row=paid_row_idx, column=col_idx)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = paid_fill
        c.border = thin_border
        if col_idx >= 4 and col_idx < 4 + len(family_ids) + 1:
            c.number_format = "$#,##0.00"
            c.alignment = Alignment(horizontal="right")
    current_row += 1

    # Summary Row 3: Net Balance (Bank Status = Paid - Owed)
    row_bal = ["Net Balance (Bank Status)", "-", "-", "$0.00"]
    for idx, fid in enumerate(family_ids):
        col_letter = get_column_letter(5 + idx)
        row_bal.append(f"={col_letter}{paid_row_idx}-{col_letter}{owed_row_idx}")
    row_bal.append("Match Bank Calculation")

    ws.append(row_bal)
    bal_row_idx = current_row
    bal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    double_bottom = Border(
        top=Side(style="thin", color="1F4E78"),
        bottom=Side(style="double", color="1F4E78"),
    )

    for col_idx in range(1, len(row_bal) + 1):
        c = ws.cell(row=bal_row_idx, column=col_idx)
        c.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
        c.fill = bal_fill
        c.border = double_bottom
        if col_idx >= 4 and col_idx < 4 + len(family_ids) + 1:
            c.number_format = "$#,##0.00;($#,##0.00);$0.00"
            c.alignment = Alignment(horizontal="right")
    current_row += 1

    # 5. Adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format and "$" in cell.number_format:
                val_str += "    "
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
